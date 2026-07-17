"""Tests unitaires — import/consultation du registre BE (`app.services.registre_be`)."""

from __future__ import annotations

import io
from typing import Any

import pytest
from flask import Flask

from app.services import registre_be as svc


def _build_registre_xlsx(rows: list[tuple[Any, ...]]) -> io.BytesIO:
    """Construit un classeur minimal imitant la structure du registre BE réel.

    ``rows`` : tuples ``(client, n_commande, destinataire, repere, type_appareil,
    nb, item, numero_affaire_libelle, annee)`` — optionnellement prolongés de
    ``(certification, categorie, module)`` (colonnes R/S/T) — insérés à partir
    de la ligne 5, colonnes 4 à 20 (comme le fichier source).
    """
    import openpyxl  # noqa: PLC0415

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BE_EF_09_A_Registre"
    ws.cell(row=3, column=4, value="CLIENT")  # en-têtes minimales, non lues par le parseur

    for i, row in enumerate(rows):
        client, n_commande, destinataire, repere, type_appareil, nb, item, libelle, annee = row[:9]
        certification, categorie, module = (row[9:12] + (None,) * 3)[:3]
        r = 5 + i
        ws.cell(row=r, column=4, value=client)
        ws.cell(row=r, column=5, value=n_commande)
        ws.cell(row=r, column=6, value=destinataire)
        ws.cell(row=r, column=7, value=repere)
        ws.cell(row=r, column=8, value=type_appareil)
        ws.cell(row=r, column=9, value=nb)
        ws.cell(row=r, column=12, value=item)
        ws.cell(row=r, column=14, value=libelle)
        ws.cell(row=r, column=16, value=annee)
        ws.cell(row=r, column=18, value=certification)
        ws.cell(row=r, column=19, value=categorie)
        ws.cell(row=r, column=20, value=module)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


_SAMPLE_ROWS = [
    (
        "Chantiers de l'Atlantique",
        "AEQU-001",
        "St NAZAIRE",
        "Navire Y34",
        "W2 03-01-12",
        7,
        8975,
        "BN0811 - RM11721",
        2026,
        "DESP 2014/68/UE",
        "IV",
        "H1",
    ),
    (
        "Chantiers de l'Atlantique",
        "AEQU-002",
        "St NAZAIRE",
        "Navire Y34",
        "W2 04-01-12",
        6,
        8976,
        "BN0811 - RM11722",
        2026,
        "STAMP U",
        None,
        None,
    ),
    (
        "TotalEnergies",
        "CMD-042",
        "GONFREVILLE",
        "322TK4131",
        "H1 06-01-72",
        1,
        9001,
        "BP0042 - TK4131",
        2025,
        "DESP + STAMP U",
        "II",
        "D1",
    ),
    (
        "Client sans affaire",
        None,
        None,
        None,
        None,
        None,
        4242,
        "RM99999 (ancienne numérotation)",
        2010,
    ),
    (
        "Client sans item",
        None,
        None,
        None,
        None,
        None,
        None,
        "BN0999 - XYZ",
        2026,
    ),
]


class TestParseRegistreExcel:
    def test_extrait_les_lignes_exploitables(self) -> None:
        buf = _build_registre_xlsx(_SAMPLE_ROWS)
        rows = svc.parse_registre_excel(buf)
        # Seules les 3 lignes avec un n° d'affaire BN/BP + item numérique sont gardées.
        assert len(rows) == 3
        assert {r.numero_affaire for r in rows} == {"BN0811", "BP0042"}

    def test_numero_affaire_et_item_formates(self) -> None:
        buf = _build_registre_xlsx(_SAMPLE_ROWS)
        rows = {r.item: r for r in svc.parse_registre_excel(buf)}
        assert rows["8975"].numero_affaire == "BN0811"
        assert rows["8975"].client_nom == "Chantiers de l'Atlantique"
        assert rows["8975"].nombre == 7
        assert rows["8975"].annee == 2026
        assert rows["9001"].numero_affaire == "BP0042"

    def test_feuille_absente_leve_key_error(self) -> None:
        import openpyxl  # noqa: PLC0415

        wb = openpyxl.Workbook()
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        with pytest.raises(KeyError):
            svc.parse_registre_excel(buf)


class TestImportRegistreBe:
    def test_import_cree_les_lignes(self, app: Flask) -> None:
        with app.app_context():
            buf = _build_registre_xlsx(_SAMPLE_ROWS)
            stats = svc.import_registre_be(buf)
            assert stats.lignes_lues == 3
            assert stats.crees == 3
            assert stats.mis_a_jour == 0

    def test_import_est_idempotent(self, app: Flask) -> None:
        with app.app_context():
            svc.import_registre_be(_build_registre_xlsx(_SAMPLE_ROWS))
            stats = svc.import_registre_be(_build_registre_xlsx(_SAMPLE_ROWS))
            assert stats.crees == 0
            assert stats.mis_a_jour == 0

    def test_reimport_met_a_jour_les_champs_modifies(self, app: Flask) -> None:
        with app.app_context():
            svc.import_registre_be(_build_registre_xlsx(_SAMPLE_ROWS))
            modifie = list(_SAMPLE_ROWS)
            modifie[0] = (
                "Chantiers de l'Atlantique — nouveau nom",
                "AEQU-001",
                "St NAZAIRE",
                "Navire Y34",
                "W2 03-01-12",
                7,
                8975,
                "BN0811 - RM11721",
                2026,
            )
            stats = svc.import_registre_be(_build_registre_xlsx(modifie))
            assert stats.crees == 0
            assert stats.mis_a_jour == 1


def _row_certification(
    item: int, certification: Any, categorie: Any = None, module: Any = None
) -> tuple[Any, ...]:
    """Ligne minimale valide avec les colonnes R/S/T paramétrables."""
    return (
        "Client",
        None,
        None,
        None,
        None,
        1,
        item,
        f"BN0900 - RM{item}",
        2026,
        certification,
        categorie,
        module,
    )


class TestReglementation:
    """Colonnes R/S/T du registre : certification, catégorie, module (V1.2 Lot 0)."""

    def test_flags_desp_et_stamp_u_sur_variantes_reelles(self) -> None:
        # Libellés réellement observés dans le fichier BE (fautes de frappe incluses).
        cas = [
            ("DESP 2014/68/UE", True, False),
            ("DESP2014/68/UE", True, False),  # sans espace
            ("DESP 2014/98/UE", True, False),  # faute de frappe réelle
            ("DESP (sans CE)", True, False),
            ("STAMP U", False, True),
            ("STAMP U-2", False, True),
            ("DESP + STAMP U", True, True),
            ("DESP 2014/68/EU + STAMP U", True, True),
            ("Décret 2015/799 du 01/07/2015", False, False),
            ("DRIRE", False, False),
            (None, False, False),
        ]
        rows = [
            _row_certification(8000 + i, certification)
            for i, (certification, _, _) in enumerate(cas)
        ]
        parsed = {r.item: r for r in svc.parse_registre_excel(_build_registre_xlsx(rows))}
        for i, (certification, desp, stamp_u) in enumerate(cas):
            row = parsed[f"{8000 + i}"]
            assert row.desp is desp, f"desp pour {certification!r}"
            assert row.stamp_u is stamp_u, f"stamp_u pour {certification!r}"
            assert row.certification_brute == certification

    def test_categorie_et_module_stockes(self, app: Flask) -> None:
        with app.app_context():
            svc.import_registre_be(_build_registre_xlsx(_SAMPLE_ROWS))
            item = svc.get_item("BN0811", "8975")
            assert item is not None
            assert item.certification_brute == "DESP 2014/68/UE"
            assert item.desp is True
            assert item.stamp_u is False
            assert item.categorie_risque == "IV"
            assert item.module_evaluation == "H1"

            stamp_seul = svc.get_item("BN0811", "8976")
            assert stamp_seul is not None
            assert stamp_seul.desp is False
            assert stamp_seul.stamp_u is True
            assert stamp_seul.categorie_risque is None
            assert stamp_seul.module_evaluation is None

    def test_categorie_numerique_convertie_en_texte(self) -> None:
        # La colonne S contient parfois 3.3 / 4.3 en numérique dans Excel.
        rows = [_row_certification(8100, "DESP", 4.3, "A")]
        (parsed,) = svc.parse_registre_excel(_build_registre_xlsx(rows))
        assert parsed.categorie_risque == "4.3"

    def test_reimport_met_a_jour_la_reglementation(self, app: Flask) -> None:
        with app.app_context():
            svc.import_registre_be(_build_registre_xlsx(_SAMPLE_ROWS))
            modifie = list(_SAMPLE_ROWS)
            modifie[0] = modifie[0][:11] + ("G",)  # module H1 → G
            stats = svc.import_registre_be(_build_registre_xlsx(modifie))
            assert stats.crees == 0
            assert stats.mis_a_jour == 1
            item = svc.get_item("BN0811", "8975")
            assert item is not None
            assert item.module_evaluation == "G"


class TestConsultation:
    def test_list_numeros_affaire(self, app: Flask) -> None:
        with app.app_context():
            svc.import_registre_be(_build_registre_xlsx(_SAMPLE_ROWS))
            numeros = svc.list_numeros_affaire()
            assert {r["numero_affaire"] for r in numeros} == {"BN0811", "BP0042"}
            bn0811 = next(r for r in numeros if r["numero_affaire"] == "BN0811")
            assert bn0811["nb_items"] == 2

    def test_list_numeros_affaire_filtre(self, app: Flask) -> None:
        with app.app_context():
            svc.import_registre_be(_build_registre_xlsx(_SAMPLE_ROWS))
            numeros = svc.list_numeros_affaire(q="BP")
            assert [r["numero_affaire"] for r in numeros] == ["BP0042"]

    def test_list_items_for_numero(self, app: Flask) -> None:
        with app.app_context():
            svc.import_registre_be(_build_registre_xlsx(_SAMPLE_ROWS))
            items = svc.list_items_for_numero("BN0811")
            assert [i.item for i in items] == ["8975", "8976"]

    def test_get_item(self, app: Flask) -> None:
        with app.app_context():
            svc.import_registre_be(_build_registre_xlsx(_SAMPLE_ROWS))
            item = svc.get_item("BN0811", "8975")
            assert item is not None
            assert item.client_nom == "Chantiers de l'Atlantique"
            assert svc.get_item("BN0811", "0000") is None
