"""Import et consultation du registre général de commande BE (Bureau d'Études).

Le fichier source (``Registre general commande BE.xlsx``) est un classeur
Excel entretenu à la main par le BE, avec un historique de plusieurs milliers
de lignes. Seules les lignes récentes portent un n° d'affaire au format
``BN``/``BP`` + 4 chiffres dans la colonne « N°PLAN »/« CODE » combinée
(colonne 14 de la feuille ``BE_EF_09_A_Registre``) — les lignes plus
anciennes (numérotation historique) sont ignorées par l'import.

Mapping colonnes (feuille ``BE_EF_09_A_Registre``, en-têtes lignes 3-4,
données à partir de la ligne 5) :

    col 4  CLIENT             → ``client_nom``
    col 5  N° Commande        → ``references_client``
    col 6  DESTINATAIRE       → ``destinataire``
    col 7  REPERE N°          → ``repere_client``
    col 8  TYPE APPAREIL      → ``type_appareil``
    col 9  Nb                 → ``nombre``
    col 12 N°PLAN              → ``item`` (4 chiffres, identifiant de ligne unique)
    col 14 (combiné)           → n° d'affaire (``BN0811 - RM11721`` par ex.)
    col 16 ANNEE               → ``annee``
    col 18 CERTIFICATION       → ``certification_brute`` + flags ``desp``/``stamp_u``
    col 19 CAT                 → ``categorie_risque`` (I, II, III, IV, 4.3…)
    col 20 MODULE              → ``module_evaluation`` (A, D1, E1, G, H, H1…)

La colonne CERTIFICATION est un texte libre historique (36 libellés distincts
constatés) : elle est conservée brute, et deux drapeaux en sont dérivés —
``desp`` si le texte contient « DESP » (toutes variantes/fautes de frappe
incluses : « DESP2014/68/UE », « DESP 2014/98/UE »…) et ``stamp_u`` s'il
contient « STAMP U » (couvre « STAMP U-2 », « DESP + STAMP U »…).

``import_registre_be`` est idempotent (upsert par clé naturelle
``numero_affaire`` + ``item``), sur le modèle de ``flask seed``.

Le fichier étant souvent ouvert dans Excel (verrou Windows/OneDrive), la
lecture bascule automatiquement sur une ouverture en partage via l'API
Windows si l'ouverture normale échoue en ``PermissionError``.
"""

from __future__ import annotations

import io
import os
import re
from dataclasses import dataclass
from pathlib import Path
from typing import TYPE_CHECKING, TypedDict

if TYPE_CHECKING:
    from app.models.registre_be import RegistreBEItem

_CellValue = str | int | float | None


class NumeroAffaireRow(TypedDict):
    """Une ligne de la liste déroulante Q1 (n° d'affaire + résumé)."""

    numero_affaire: str
    client_nom: str | None
    nb_items: int

_SHEET_NAME = "BE_EF_09_A_Registre"
_FIRST_DATA_ROW = 5

_COL_CLIENT = 4
_COL_REFERENCES_CLIENT = 5
_COL_DESTINATAIRE = 6
_COL_REPERE_CLIENT = 7
_COL_TYPE_APPAREIL = 8
_COL_NOMBRE = 9
_COL_ITEM = 12
_COL_NUMERO_AFFAIRE = 14
_COL_ANNEE = 16
_COL_CERTIFICATION = 18
_COL_CATEGORIE = 19
_COL_MODULE = 20

_NUMERO_AFFAIRE_RE = re.compile(r"(BN|BP)(\d{4})(?!\d)", re.IGNORECASE)


@dataclass(frozen=True)
class RegistreRow:
    """Une ligne exploitable du registre BE, prête à être upsertée."""

    numero_affaire: str
    item: str
    client_nom: str | None
    destinataire: str | None
    repere_client: str | None
    type_appareil: str | None
    nombre: int | None
    annee: int | None
    references_client: str | None
    libelle_brut: str
    certification_brute: str | None = None
    desp: bool = False
    stamp_u: bool = False
    categorie_risque: str | None = None
    module_evaluation: str | None = None


@dataclass(frozen=True)
class ImportStats:
    """Résultat d'un import du registre BE."""

    lignes_lues: int
    lignes_ignorees: int
    crees: int
    mis_a_jour: int


def parse_registre_excel(path: str | Path) -> list[RegistreRow]:
    """Parse le classeur Excel et retourne les lignes exploitables.

    Ignore silencieusement les lignes sans n° d'affaire ``BN``/``BP`` + 4
    chiffres reconnaissable, ou sans n° d'item numérique en colonne N°PLAN.

    Args:
        path: Chemin vers le fichier ``.xlsx`` du registre BE.

    Returns:
        La liste des lignes exploitables (une par item d'affaire).

    Raises:
        KeyError: Si la feuille ``BE_EF_09_A_Registre`` est absente du classeur.
    """
    import openpyxl  # noqa: PLC0415

    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except PermissionError:
        # Fichier verrouillé (typiquement ouvert dans Excel / synchronisé
        # OneDrive) : relecture en partage via l'API Windows.
        wb = openpyxl.load_workbook(
            _read_locked_file(path), data_only=True, read_only=True
        )
    if _SHEET_NAME not in wb.sheetnames:
        raise KeyError(f"Feuille {_SHEET_NAME!r} introuvable dans {path!r}.")
    ws = wb[_SHEET_NAME]

    rows: list[RegistreRow] = []
    for row in ws.iter_rows(min_row=_FIRST_DATA_ROW):
        cells = [c.value for c in row]

        def cell(col: int, cells: list[_CellValue] = cells) -> _CellValue:
            return cells[col - 1] if col - 1 < len(cells) else None

        libelle_brut = cell(_COL_NUMERO_AFFAIRE)
        if not isinstance(libelle_brut, str):
            continue
        match = _NUMERO_AFFAIRE_RE.search(libelle_brut)
        if match is None:
            continue

        item_raw = cell(_COL_ITEM)
        if not isinstance(item_raw, (int, float)):
            continue
        item = f"{int(item_raw):04d}"

        certification = _as_str(cell(_COL_CERTIFICATION))
        certification_upper = (certification or "").upper()

        rows.append(
            RegistreRow(
                numero_affaire=f"{match.group(1).upper()}{match.group(2)}",
                item=item,
                client_nom=_as_str(cell(_COL_CLIENT)),
                destinataire=_as_str(cell(_COL_DESTINATAIRE)),
                repere_client=_as_str(cell(_COL_REPERE_CLIENT)),
                type_appareil=_as_str(cell(_COL_TYPE_APPAREIL)),
                nombre=_as_int(cell(_COL_NOMBRE)),
                annee=_as_int(cell(_COL_ANNEE)),
                references_client=_as_str(cell(_COL_REFERENCES_CLIENT)),
                libelle_brut=libelle_brut.strip(),
                certification_brute=certification,
                desp="DESP" in certification_upper,
                stamp_u="STAMP U" in certification_upper,
                categorie_risque=_as_str(cell(_COL_CATEGORIE)),
                module_evaluation=_as_str(cell(_COL_MODULE)),
            )
        )
    return rows


def _read_locked_file(path: str | Path) -> io.BytesIO:
    """Lit un fichier verrouillé par un autre processus (Excel) en partage.

    L'ouverture standard de CPython échoue en ``PermissionError`` quand Excel
    détient le fichier ; l'API Windows ``CreateFileW`` avec les trois flags
    ``FILE_SHARE_*`` permet une lecture concurrente sans copie préalable.

    Args:
        path: Chemin du fichier verrouillé.

    Returns:
        Le contenu complet du fichier dans un tampon mémoire.

    Raises:
        PermissionError: Si la lecture partagée échoue aussi (ou hors Windows).
    """
    if os.name != "nt":  # pragma: no cover — verrou Excel = cas Windows
        raise PermissionError(
            f"{path} est verrouillé — fermez le fichier dans Excel puis relancez."
        )

    import ctypes  # noqa: PLC0415
    import msvcrt  # noqa: PLC0415

    generic_read = 0x80000000
    file_share_all = 0x1 | 0x2 | 0x4  # READ | WRITE | DELETE
    open_existing = 3

    handle = ctypes.windll.kernel32.CreateFileW(
        str(Path(path).resolve()),
        generic_read,
        file_share_all,
        None,
        open_existing,
        0,
        None,
    )
    if handle in (-1, 0xFFFFFFFFFFFFFFFF):
        raise PermissionError(
            f"{path} est verrouillé et la lecture partagée a échoué — "
            "fermez le fichier dans Excel puis relancez."
        )
    fd = msvcrt.open_osfhandle(handle, os.O_RDONLY)
    with os.fdopen(fd, "rb") as fh:
        return io.BytesIO(fh.read())


def import_registre_be(path: str | Path) -> ImportStats:
    """Importe (upsert idempotent) le registre BE dans ``registre_be_items``.

    Args:
        path: Chemin vers le fichier ``.xlsx`` du registre BE.

    Returns:
        Les statistiques de l'import (lignes lues/ignorées, créées/mises à jour).
    """
    from app.extensions import db  # noqa: PLC0415
    from app.models.registre_be import RegistreBEItem  # noqa: PLC0415

    parsed = parse_registre_excel(path)

    existing = {(r.numero_affaire, r.item): r for r in db.session.query(RegistreBEItem).all()}

    crees = 0
    mis_a_jour = 0
    for row in parsed:
        key = (row.numero_affaire, row.item)
        record = existing.get(key)
        if record is None:
            db.session.add(
                RegistreBEItem(
                    numero_affaire=row.numero_affaire,
                    item=row.item,
                    client_nom=row.client_nom,
                    destinataire=row.destinataire,
                    repere_client=row.repere_client,
                    type_appareil=row.type_appareil,
                    nombre=row.nombre,
                    annee=row.annee,
                    references_client=row.references_client,
                    libelle_brut=row.libelle_brut,
                    certification_brute=row.certification_brute,
                    desp=row.desp,
                    stamp_u=row.stamp_u,
                    categorie_risque=row.categorie_risque,
                    module_evaluation=row.module_evaluation,
                )
            )
            crees += 1
        else:
            changed = False
            for field in (
                "client_nom",
                "destinataire",
                "repere_client",
                "type_appareil",
                "nombre",
                "annee",
                "references_client",
                "libelle_brut",
                "certification_brute",
                "desp",
                "stamp_u",
                "categorie_risque",
                "module_evaluation",
            ):
                value = getattr(row, field)
                if getattr(record, field) != value:
                    setattr(record, field, value)
                    changed = True
            if changed:
                mis_a_jour += 1

    db.session.commit()
    return ImportStats(
        lignes_lues=len(parsed),
        lignes_ignorees=0,
        crees=crees,
        mis_a_jour=mis_a_jour,
    )


# ── Consultation (wizard Q1) ─────────────────────────────────────────────


def list_numeros_affaire(q: str | None = None) -> list[NumeroAffaireRow]:
    """Liste les n° d'affaire distincts du registre pour la liste déroulante Q1.

    Args:
        q: Filtre optionnel (recherche partielle sur le n° d'affaire).

    Returns:
        Liste de ``{"numero_affaire": ..., "client_nom": ..., "nb_items": ...}``
        triée par n° d'affaire.
    """
    from sqlalchemy import func  # noqa: PLC0415

    from app.extensions import db  # noqa: PLC0415
    from app.models.registre_be import RegistreBEItem  # noqa: PLC0415

    stmt = db.session.query(
        RegistreBEItem.numero_affaire,
        func.max(RegistreBEItem.client_nom),
        func.count(RegistreBEItem.id),
    ).group_by(RegistreBEItem.numero_affaire)

    if q:
        stmt = stmt.filter(RegistreBEItem.numero_affaire.ilike(f"%{q}%"))

    return [
        {"numero_affaire": numero, "client_nom": client, "nb_items": nb}
        for numero, client, nb in stmt.order_by(RegistreBEItem.numero_affaire).all()
    ]


def list_items_for_numero(numero_affaire: str) -> list[RegistreBEItem]:
    """Liste les items disponibles pour un n° d'affaire donné (dropdown Q1)."""
    from app.extensions import db  # noqa: PLC0415
    from app.models.registre_be import RegistreBEItem  # noqa: PLC0415

    return (
        db.session.query(RegistreBEItem)
        .filter(RegistreBEItem.numero_affaire == numero_affaire)
        .order_by(RegistreBEItem.item)
        .all()
    )


def get_item(numero_affaire: str, item: str) -> RegistreBEItem | None:
    """Résout une ligne du registre par (n° affaire, n° item)."""
    from app.extensions import db  # noqa: PLC0415
    from app.models.registre_be import RegistreBEItem  # noqa: PLC0415

    return (
        db.session.query(RegistreBEItem)
        .filter(
            RegistreBEItem.numero_affaire == numero_affaire,
            RegistreBEItem.item == item,
        )
        .first()
    )


def _as_str(value: _CellValue) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _as_int(value: _CellValue) -> int | None:
    if isinstance(value, (int, float)):
        return int(value)
    return None
